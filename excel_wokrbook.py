import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from Errors import throw

FONT_COLOR = Font(color="FFFFFF")
WRAP_TEXT = Alignment(vertical='top', wrapText=True)
FONT_COLOR_HEADER = Font(color="000000", bold=True)
BORDER_RIGHT = Border(right=Side(style="thick", color="000000"))
BORDER_TITTLE = Border(bottom=Side(style="thick", color="000000"),
                       right=Side(style="thick", color="000000"))
FONT_COLOR_WHITE_BOLD = Font(color="FFFFFF", bold=True)
FONT_COLOR_WHITE = Font(color="FFFFFF")
FONT_COLOR_BLACK_BOLD = Font(color="000000", bold=True)
BORDER_COMPLETE_BLACK = Border(top=Side(style="thin", color="000000"),
                         bottom=Side(style="thin", color="000000"),
                         right=Side(style="thin", color="000000"),
                         left=Side(style="thin", color="000000"))
BORDER_COMPLETE_WHITE = Border(top=Side(style="thin", color="FFFFFF"),
                         bottom=Side(style="thin", color="FFFFFF"),
                         right=Side(style="thin", color="FFFFFF"),
                         left=Side(style="thin", color="FFFFFF"))
GREEN_FILL = PatternFill(fill_type='solid', fgColor='5BA92C')
ALIGNMENT_HORIZONTAL_CENTER = Alignment(horizontal='center')
ALIGNMENT_CENTER_CENTER = Alignment(horizontal='center', vertical='center')
ALIGNMENT_WRAP_TEXT = Alignment(wrapText=True)
COLORSCALE = ColorScaleRule(start_type='percentile', start_value=0, start_color='FFC0B7',
                            end_type='percentile', end_value=100, end_color='BFEFAD')
SIDE_BLACK_MEDIUM = Side(style="medium", color="000000")
SIDE_BLACK_THIN = Side(style="thin", color="000000")
SIDE_WHITE_THIN = Side(style="thin", color="FFFFFF")

class ExcelWorkbook:
    def __init__(self, path: str, contexts: list[dict], all_policies_content: pd.DataFrame) -> None:
        self.path = path
        self.contexts = contexts
        self.all_policies_content = all_policies_content

        self.colors_pairs = [
            {
                'color' : 'grey',
                'hex1' : 'D9D9D9',
                'hex2' : 'BFBFBF'
            },
            {
                'color' : 'green',
                'hex1' : 'DFF8D6',
                'hex2' : 'BFEFAD'
            },
            {
                'color' : 'blue',
                'hex1' : '99D9FF',
                'hex2' : '66C6FF'
            },
            {
                'color' : 'red',
                'hex1' : 'FC4868',
                'hex2' : 'FF8CA0'
            },
            {
                'color' : 'purple',
                'hex1' : 'AD4CFF',
                'hex2' : 'D19CFD'
            },
            {
                'color' : 'orange',
                'hex1' : 'FFD045',
                'hex2' : 'FDE28F'
            }
        ]

        if not self.path.endswith('.xlsx'):
            self.path += '.xlsx'

        self.all_policies_content['Operator'] = self.all_policies_content['Operator'].replace('=|0','=CONCATENATE("=|0")')

        self.workbook = Workbook()
        self.save()
        self.workbook = load_workbook(self.path)
        self.create_dashboard_sheet()
        self.create_contexts_sheet()
        self.create_all_policies_sheet()
        self.create_contexts_sheets()
        self.save()

        self.append_all_policies()
        self.append_contexts_data()
        self.construct_global_information_header()
        self.construct_global_information_rows()
        self.construct_contexts_columns()
        self.construct_contexts_table()
        self.save()

        self.construct_dashboard_sheet()
        self.construct_workshop_dashboard_sheet()

    def save(self):
        """Saves the excel file
        """
        try:
            self.workbook.save(self.path)
        except:
            throw('An error occured while saving the Excel file, the name might be the cause.', 'high')

    def create_dashboard_sheet(self):
        """Create the dashboard sheet
        """
        sheet = self.workbook.active
        sheet.title = 'Dashboards'
        for row in sheet.iter_rows(max_row=160, max_col=50):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="FFFFFF")
        worksheet_dashboards_workshops = self.workbook.create_sheet(index=1, title='Dashboards - Workshops')
        for row in worksheet_dashboards_workshops.iter_rows(max_row=55, max_col=71):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="FFFFFF")
    
    def create_contexts_sheet(self):
        """Create the Contexts sheet
        """
        self.workbook.create_sheet(index=2, title='Contexts')
    
    def create_all_policies_sheet(self):
        """Create the All_Policies sheet
        """
        worksheet = self.workbook.create_sheet(index=3, title='All-Policies')
        worksheet.sheet_properties.tabColor = 'FFDFDB'

    def create_contexts_sheets(self):
        """Create different contexts sheets
        """
        sheet_index = 4
        context_number = 1
        colors_index = 0
        colors_len = len(self.colors_pairs)
        for context in self.contexts:
            worksheet = self.workbook.create_sheet(index=sheet_index, title=f'{context_number}-{context["Name"]}-extract')
            worksheet.sheet_properties.tabColor = self.colors_pairs[colors_index]['hex1']
            sheet_index+=1
            worksheet = self.workbook.create_sheet(index=sheet_index, title=f'{context_number}-{context["Name"]}-log-reported')
            worksheet.sheet_properties.tabColor = self.colors_pairs[colors_index]['hex2']
            sheet_index+=1
            worksheet = self.workbook.create_sheet(index=sheet_index, title=f'{context_number}-{context["Name"]}-finding-list')
            sheet_index+=1

            context_number+=1

            colors_index+=1
            # restart color if all colors where used
            if colors_index > colors_len:
                colors_index = 0
                
    def append_all_policies(self):
        """This will add all_policies data to corresponding sheet
        """
        # Add all_policies data
        with pd.ExcelWriter(self.path,
                        mode='a',
                        engine='openpyxl',
                        if_sheet_exists='overlay'
        ) as writer:
            self.all_policies_content.to_excel(writer, sheet_name='All-Policies', index=False)

        # Format all_policies sheet
        self.workbook = load_workbook(self.path)
        worksheet = self.workbook['All-Policies']
        table = Table(displayName='All_policies',
                    ref=f'A1:{get_column_letter(self.all_policies_content.shape[1])}{len(self.all_policies_content)+1}')
        style = TableStyleInfo(name="TableStyleMedium11", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        worksheet.add_table(table)
        for cells in worksheet["1:1"]:
            cells.font = FONT_COLOR
        for columns in range(worksheet.min_column, worksheet.max_column + 1):
            worksheet.column_dimensions[get_column_letter(columns)].width = 25
        self.save()

    def append_contexts_data(self):
        """This will add contexts data to corresponding sheet
        """
        context_number = 1
        for context in self.contexts:
            # Append LOG
            with pd.ExcelWriter(self.path,
                        mode='a',
                        engine='openpyxl',
                        if_sheet_exists='overlay'
            ) as writer:
                context['Log'].to_excel(writer, sheet_name=f'{context_number}-{context["Name"]}-log-reported', index=False)

            # Format LOG
            self.workbook = load_workbook(self.path)
            worksheet = self.workbook[f'{context_number}-{context["Name"]}-log-reported']
            table = Table(displayName=f'context{context_number}_log', ref=f'A1:{get_column_letter(context["Log"].shape[1])}{len(context["Log"])+1}')
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            worksheet.add_table(table)
            for cells in worksheet["1:1"]:
                cells.font = FONT_COLOR
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = WRAP_TEXT
            for columns in range(worksheet.min_column, worksheet.max_column + 1):
                worksheet.column_dimensions[get_column_letter(columns)].width = 100
            self.save()

            # Append FINDING_LIST
            with pd.ExcelWriter(self.path,
                        mode='a',
                        engine='openpyxl',
                        if_sheet_exists='replace'
            ) as writer:
                context['FindingList'].to_excel(writer, sheet_name=f'{context_number}-{context["Name"]}-finding-list', index=False)
            
            # Format FINDING_LIST
            self.workbook = load_workbook(self.path)
            worksheet = self.workbook[f'{context_number}-{context["Name"]}-finding-list']
            table = Table(displayName=f'finding_list_context{context_number}',
                        ref=f'A1:{get_column_letter(context["FindingList"].shape[1])}{len(context["FindingList"])+1}')
            style = TableStyleInfo(name="TableStyleMedium11", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            worksheet.add_table(table)
            for cells in worksheet["1:1"]:
                cells.font = FONT_COLOR
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = WRAP_TEXT
            for columns in range(worksheet.min_column, worksheet.max_column + 1):
                worksheet.column_dimensions[get_column_letter(columns)].width = 50
            self.save()

            # Append EXTRACT
            with pd.ExcelWriter(self.path,
                        mode='a',
                        engine='openpyxl',
                        if_sheet_exists='overlay'
            ) as writer:
                context['Extract'].to_excel(writer, sheet_name=f'{context_number}-{context["Name"]}-extract', index=False)

            # Format EXTRACT
            self.workbook = load_workbook(self.path)
            worksheet = self.workbook[f'{context_number}-{context["Name"]}-extract']
            table = Table(displayName=f'context{context_number}_reported_file_finding_list',
                        ref=f'A1:{get_column_letter(context["Extract"].shape[1])}{len(context["Extract"])+1}')
            style = TableStyleInfo(name="TableStyleMedium8", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            worksheet.add_table(table)
            for cells in worksheet["1:1"]:
                cells.font = FONT_COLOR
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = WRAP_TEXT
            for columns in range(worksheet.min_column, worksheet.max_column + 1):
                worksheet.column_dimensions[get_column_letter(columns)].width = 50
            self.save()
            
            context_number+=1

    def construct_global_information_header(self):
        """
        Table Header Global Information
        """
        worksheet = self.workbook['Contexts']
        worksheet['A1'] = 'Global Information'
        worksheet['A1'].alignment = Alignment(horizontal="center")
        worksheet['A1'].font = FONT_COLOR_HEADER
        worksheet['A1'].border = BORDER_TITTLE
        worksheet.merge_cells('A1:H1')
        worksheet['A2'] = 'ID'
        worksheet['B2'] = 'Line'
        worksheet['C2'] = 'Category'
        worksheet['D2'] = 'Name'
        worksheet['E2'] = 'CIS Level'
        worksheet['F2'] = 'Severity from original file'
        worksheet['G2'] = 'Default Value'
        worksheet['H2'] = 'Recommended Value'
        worksheet['H2'].border = BORDER_RIGHT

    def construct_global_information_rows(self):
        """
        Create all columns (ID, Line, etc.) for the global information
        """
        worksheet_all_policies = self.workbook['All-Policies']
        worksheet_contexts = self.workbook['Contexts']
        row_all_policies = worksheet_all_policies.max_row - 1
        value_columns_global = []
        for row_number in range(row_all_policies):
            value_columns_global.append([
                f"='All-Policies'!A{row_number + 2}",#ID
                "=ROW(INDEX(All_policies[Rationale],MATCH([[#This Row],[Name]],All_policies[Name],0)))-1",#Line
                "=INDEX(All_policies[Category],MATCH([[#This Row],[Name]],All_policies[Name],0))",#Category
                f"='All-Policies'!C{row_number + 2}",#Name
                "=INDEX(All_policies[Level],MATCH([[#This Row],[Name]],All_policies[Name],0))",#Level
                "=INDEX(All_policies[Severity],MATCH([[#This Row],[Name]],All_policies[Name],0))",#Severity
                "=CONCATENATE(INDEX(All_policies[DefaultValue],MATCH([[#This Row],[Name]],All_policies[Name],0)))",#DefaultValue
                "=CONCATENATE(INDEX(All_policies[RecommendedValue],MATCH([[#This Row],[Name]],All_policies[Name],0)))"#RecommendedValue
            ])
            for column_number in range(8):
                worksheet_contexts.cell(row=row_number+3,column=column_number+1).value =\
                value_columns_global[row_number][column_number]
                worksheet_contexts.cell(row=row_number+3,column=column_number+1).font =\
                    Font(size=11, color="000000", bold=True)
                worksheet_contexts.cell(row=row_number+3,column=column_number+1).fill =\
                    PatternFill(fill_type='solid', fgColor='DDDDDD')
                worksheet_contexts.cell(row=row_number+3,column=8).border = BORDER_RIGHT

    def construct_contexts_columns(self):
        """This function creates contexts column and fill them
        """
        worksheet = self.workbook['Contexts']
        context_number = 1
        start_line = 9
        end_line = start_line + 8
        colors_index = 0
        colors_len = len(self.colors_pairs)
        for context in self.contexts:
            self.construct_context_header(worksheet, context['Name'], context_number, start_line, end_line)
            self.construct_context_rows(context_number, start_line, end_line, self.colors_pairs[colors_index]['hex2'])
            context_number+=1
            colors_index+=1
            if context_number <= len(self.contexts):
                start_line+=9
                end_line+=9
            # restart color if all colors where used
            if colors_index > colors_len:
                colors_index = 0
        self.construct_workshops_header(worksheet, end_line+1)
        self.construct_workshops_rows(end_line+1)

    def construct_context_header(self, worksheet, nom_context,
                             num_context, number_column_start, number_column_end):
        """
        Table Header for each context
        """
        letter_column_start = get_column_letter(number_column_start)
        letter_column_end = get_column_letter(number_column_end)
        worksheet[letter_column_start + '1'] = 'Context-' + str(num_context) + '(' + nom_context + ')'
        worksheet[letter_column_start + '1'].alignment = Alignment(horizontal="center")
        worksheet[letter_column_start + '1'].font = FONT_COLOR_HEADER
        worksheet[letter_column_start + '1'].border = BORDER_TITTLE
        worksheet.merge_cells(letter_column_start + '1:' + letter_column_end + '1')
        list_columns = []
        for columns in range(number_column_start, number_column_end + 1):
            column_letter = get_column_letter(columns)
            list_columns.append(column_letter)
        worksheet[list_columns[0] + '2'] = 'ID Context' + str(num_context)
        worksheet[list_columns[1] + '2'] = 'Context' + str(num_context) + ' - Operator'
        worksheet[list_columns[2] + '2'] = 'Context' + str(num_context) + ' - Computed Severity'
        worksheet[list_columns[3] + '2'] = 'Context' + str(num_context) + ' - Manual Severity'
        worksheet[list_columns[4] + '2'] = 'Context' + str(num_context) + ' - Result'
        worksheet[list_columns[5] + '2'] = 'Context' + str(num_context) + ' - ComputedResult'
        worksheet[list_columns[6] + '2'] = 'Context' + str(num_context) + ' - Fixed Value'
        worksheet[list_columns[7] + '2'] = 'Context' + str(num_context) + ' - Computed Value'
        worksheet[list_columns[8] + '2'] = 'Context' + str(num_context) + ' - isRecValue'
        worksheet[list_columns[8] + '2'].border = BORDER_RIGHT

    def construct_context_rows(self, numero_context, first_column, last_column, color_context):
        """
        Create all columns (ID, Result, Severity, etc.) for each context
        """
        worksheet_all_policies = self.workbook['All-Policies']
        worksheet_contexts = self.workbook['Contexts']
        row_all_policies = worksheet_all_policies.max_row - 1
        value_columns_context = []
        for row_number in range(row_all_policies):
            value_columns_context.append([f"=INDEX(finding_list_context{numero_context}[ID],MATCH([Name],finding_list_context{numero_context}[Name],0))",
                                f"=INDEX(finding_list_context{numero_context}[Operator],MATCH([Name],finding_list_context{numero_context}[Name],0))",
                                f'=IF(ISERROR([[#This Row],[ID context{numero_context}]]),"Not Applicable",IF(ISBLANK([[#This Row],[context{numero_context} - Manual Severity]]),IFERROR(_xlfn.SWITCH([[#This Row],[context{numero_context} - Operator]],">=",IF(_xlfn.NUMBERVALUE([[#This Row],[context{numero_context} - ComputedResult]])>=_xlfn.NUMBERVALUE([[#This Row],[Recommended Value]]),"Passed",[[#This Row],[Severity from original file]]),"<=!0",IF(AND(_xlfn.NUMBERVALUE([[#This Row],[context{numero_context} - ComputedResult]])<=_xlfn.NUMBERVALUE([[#This Row],[Recommended Value]]),[[#This Row],[context{numero_context} - ComputedResult]]<>0),"Passed",[[#This Row],[Severity from original file]]),0,IF([[#This Row],[Recommended Value]]=[[#This Row],[context{numero_context} - ComputedResult]],"Passed",[[#This Row],[Severity from original file]]),"!=",IF([[#This Row],[Recommended Value]]<>[[#This Row],[context{numero_context} - ComputedResult]],"Passed",[[#This Row],[Severity from original file]]),"=",IF([[#This Row],[Recommended Value]]=[[#This Row],[context{numero_context} - ComputedResult]],"Passed",[[#This Row],[Severity from original file]]),"<=",IF(_xlfn.NUMBERVALUE([[#This Row],[context{numero_context} - ComputedResult]])<=_xlfn.NUMBERVALUE([[#This Row],[Recommended Value]]),"Passed",[[#This Row],[Severity from original file]]),"contains",IF(SEARCH([[#This Row],[Recommended Value]],[[#This Row],[context{numero_context} - Result]]),"Passed",[[#This Row],[Severity from original file]])),[[#This Row],[Severity from original file]]),[[#This Row],[context{numero_context} - Manual Severity]]))',
                                '',
                                f"=CONCATENATE(INDEX(context{numero_context}_reported_file_finding_list[Result],MATCH([Name],context{numero_context}_reported_file_finding_list[Name],0)))",
                                f'=_xlfn.SWITCH([[#This Row],[context{numero_context} - Result]],"-NODATA-",[[#This Row],[Default Value]],"BUILTIN\Administrateurs","BUILTIN\Administrators",[[#This Row],[context{numero_context} - Result]])',
                                "_",
                                f'=_xlfn.SWITCH([[#This Row],[context{numero_context} - Fixed Value]],"same",[[#This Row],[context{numero_context} - ComputedResult]],"recval",[[#This Row],[Recommended Value]],"to check","to check","_","N/A",[[#This Row],[context{numero_context} - Fixed Value]])',
                                f'=_xlfn.SWITCH([[#This Row],[context{numero_context} - Fixed Value]],"_",[[#This Row],[context{numero_context} - Computed Severity]],"recval","Passed",IFERROR(_xlfn.SWITCH([[#This Row],[context{numero_context} - Operator]],">=",IF(_xlfn.NUMBERVALUE([[#This Row],[context{numero_context} - Fixed Value]])>=_xlfn.NUMBERVALUE([[#This Row],[Recommended Value]]),"Passed",[[#This Row],[context{numero_context} - Computed Severity]]),"<=!0",IF(AND(_xlfn.NUMBERVALUE([[#This Row],[context{numero_context} - Fixed Value]])<=_xlfn.NUMBERVALUE([[#This Row],[Recommended Value]]),[[#This Row],[context{numero_context} - Fixed Value]]<>0),"Passed",[[#This Row],[context{numero_context} - Computed Severity]]),0,IF([[#This Row],[Recommended Value]]=[[#This Row],[context{numero_context} - Fixed Value]],"Passed",[[#This Row],[context{numero_context} - Computed Severity]]),"!=",IF([[#This Row],[Recommended Value]]<>[[#This Row],[context{numero_context} - Fixed Value]],"Passed",[[#This Row],[context{numero_context} - Computed Severity]]),"<=",IF(_xlfn.NUMBERVALUE([[#This Row],[context{numero_context} - Fixed Value]])<=_xlfn.NUMBERVALUE([[#This Row],[Recommended Value]]),"Passed",[[#This Row],[context{numero_context} - Computed Severity]]),"contains",IF(SEARCH([[#This Row],[Recommended Value]],[[#This Row],[context{numero_context} - Fixed Value]]),"Passed",[[#This Row],[context{numero_context} - Computed Severity]])),[[#This Row],[context{numero_context} - Computed Severity]]))'])
            for column_number in range(9):
                worksheet_contexts.cell(row=row_number+3,column=column_number+first_column).value =\
                    value_columns_context[row_number][column_number]
                worksheet_contexts.cell(row=row_number+3,column=column_number+first_column).fill =\
                    PatternFill(fill_type='solid', fgColor=color_context)
                worksheet_contexts.cell(row=row_number+3,column=last_column).border = BORDER_RIGHT

    def construct_workshops_header(self, worksheet, number_column_start):
        """
        Table header for different workshops, commentary etc.
        """
        letter_column_start = get_column_letter(number_column_start)
        worksheet[letter_column_start + '1'] = 'Summary'
        worksheet[letter_column_start + '1'].alignment = Alignment(horizontal="center")
        worksheet[letter_column_start + '1'].font = FONT_COLOR_HEADER
        worksheet[letter_column_start + '1'].border = BORDER_TITTLE
        list_columns = []
        for columns in range(number_column_start, number_column_start + 11):
            column_letter = get_column_letter(columns)
            list_columns.append(column_letter)
        worksheet[list_columns[0] + '2'] = 'Workshops'
        worksheet[list_columns[1] + '2'] = 'All Passed'
        worksheet[list_columns[2] + '2'] = 'At least one "to check"'

        if len(self.contexts) == 1:
            number_of_columns = 7
        else:
            number_of_columns = 6 + len(self.contexts)
        letter_column_end = get_column_letter(number_column_start + number_of_columns)
        worksheet.merge_cells(letter_column_start + '1:' + letter_column_end + '1')

        context_index = 1
        col = 3
        all_column_name = 'Context1'
        for _ in self.contexts:
            worksheet[list_columns[col] + '2'] = f'Context{context_index}'
            if len(self.contexts) > 1:
                all_column_name+=f' & Context{context_index}'
            context_index+=1
            col+=1
        if len(self.contexts) > 1:
            worksheet[list_columns[col] + '2'] = all_column_name
            col+=1
        worksheet[list_columns[col] + '2'] = 'Actions (Coté Cryptonit)'
        worksheet[list_columns[col+1] + '2'] = 'Actions (Coté Client)'
        worksheet[list_columns[col+2] + '2'] = 'Conclusion'
        worksheet[list_columns[col+2] + '2'].border = BORDER_RIGHT

    def construct_workshops_rows(self, number_column_start):
        """
        Create all columns for the summary
        """
        worksheet_all_policies = self.workbook['All-Policies']
        worksheet_contexts = self.workbook['Contexts']
        row_all_policies = worksheet_all_policies.max_row - 1
        value_columns_global = []
        
        for row_number in range(row_all_policies):
            current_row_formulas = []
            current_row_formulas.append('_')

            formula_1 = '=AND('
            end_formula_1 = ')'

            formula_2 = '=OR('
            end_formula_2 = ')'

            context_index = 1
            for _ in self.contexts:
                formula_1+=f'[[#This Row],[Context{context_index} - Computed Severity]]="Passed"'
                formula_2+=f'[[#This Row],[Context{context_index} - Fixed Value]]="to check"'
                if len(self.contexts) > context_index:
                    formula_1+=','
                    formula_2+=','
                context_index+=1

            formula_1+=end_formula_1
            formula_2+=end_formula_2

            current_row_formulas.append(formula_1)
            current_row_formulas.append(formula_2)

            context_index = 1
            formula_3 = '=AND('
            end_formula_3 = ')'
            for _ in self.contexts:
                current_row_formulas.append(f'=IF(IFERROR([[#This Row],[ID Context{context_index}]],FALSE)=FALSE,FALSE,TRUE)')
                formula_3+=f'[[#This Row],[Context{context_index}]]'
                if len(self.contexts) > context_index:
                    formula_3+=','
                context_index+=1

            formula_3+=end_formula_3
            
            if len(self.contexts) > 1:
                current_row_formulas.append(formula_3)
            
            current_row_formulas+=['','','','']

            value_columns_global.append(current_row_formulas)
            
            default_rng = 6
            if len(self.contexts) > 1:
                default_rng+=len(self.contexts)
                

            for column_number in range(default_rng):
                worksheet_contexts.cell(row=row_number+3,
                                        column=column_number+number_column_start).value =\
                value_columns_global[row_number][column_number]
                worksheet_contexts.cell(row=row_number+3,
                                        column=column_number+number_column_start).font =\
                    Font(size=11, color="000000", bold=True)
                worksheet_contexts.cell(row=row_number+3,
                                        column=column_number+number_column_start).fill =\
                    PatternFill(fill_type='solid', fgColor='F2F2F2')
                worksheet_contexts.cell(row=row_number+3,column=8).border = BORDER_RIGHT

    def construct_contexts_table(self):
        """
        Create the table of contexts sheet
        """
        worksheet_contexts = self.workbook['Contexts']
        table = Table(displayName='table_contexts',
                    ref=f'A2:{get_column_letter(worksheet_contexts.max_column)}{worksheet_contexts.max_row}')
        style = TableStyleInfo(name="TableStyleLight15", showFirstColumn=False, showLastColumn=False,
                            showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        worksheet_contexts.add_table(table)
        for header_cells in worksheet_contexts["2:2"]:
            header_cells.font = FONT_COLOR_HEADER
            header_cells.fill = PatternFill(fill_type='solid', fgColor='9F9F9F')
        for columns in range(worksheet_contexts.min_column, worksheet_contexts.max_column + 1):
            worksheet_contexts.column_dimensions[get_column_letter(columns)].width = 15
        for row in worksheet_contexts.iter_rows(min_row=3, max_row=worksheet_contexts.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical='center', wrapText=True)

    def construct_dashboard_sheet(self):
        """
        Launch all functions to create the entire sheet
        """
        workbook = load_workbook(self.path, data_only=False)
        worksheet = workbook['Dashboards']
        for columns in range(1, 46):
            worksheet.column_dimensions[get_column_letter(columns)].width = 11
        worksheet.column_dimensions['O'].width = 45
        self.table_number_policies(worksheet)
        self.table_number_workshop(worksheet)
        self.pie_chart_policies_by_workshop_on_main_dashboard(worksheet)
        starting_row = 27
        starting_letter = 'D'
        context_index = 1
        for context in self.contexts:
            self.chart_context(worksheet, context, context_index, starting_row, f'${starting_letter}$3', f'${starting_letter}$5')
            starting_letter = get_column_letter(column_index_from_string(starting_letter) + 2)
            starting_row+=42
            context_index+=1
        # if nom_context1:
        #     chart_context(final_excel, worksheet, nom_context1, 27, 1, '$D$3', '$D$5')
        # if nom_context2:
        #     chart_context(final_excel, worksheet, nom_context2, 69, 2, '$F$3', '$F$5')
        # if nom_context3:
        #     chart_context(final_excel, worksheet, nom_context3, 111, 3, '$H$3', '$H$5')
        workbook.save(self.path)

    def table_number_policies(self, worksheet):
        """
        Create the first table with numbers of policies and categories for each context
        """
        worksheet['B3'] = 'Policies'
        worksheet['B4'] = 'Categories'
        worksheet['B5'] = 'L1'
        worksheet['B6'] = 'L2'
        for cell in range(2,7):
            worksheet[f'B{cell}'].border = BORDER_COMPLETE_BLACK
            worksheet[f'B{cell}'].alignment = ALIGNMENT_HORIZONTAL_CENTER
            worksheet.merge_cells(f'B{cell}:C{cell}')

        letter = 'D'
        context_index = 1
        for _ in self.contexts:
            worksheet[f'{letter}2'] = f'Context{context_index}'
            worksheet[f'{letter}3'] = f'=COUNTA(context{context_index}_reported_file_finding_list[ID])'
            worksheet[f'{letter}4'] = f'=COUNTA(_xlfn.UNIQUE(context{context_index}_reported_file_finding_list[Category]))'
            worksheet[f'{letter}5'] = f'=COUNTIFS(table_contexts[Context{context_index}],TRUE,table_contexts[CIS Level],"(L1)")'
            worksheet[f'{letter}6'] = f'=COUNTIFS(table_contexts[Context{context_index}],TRUE,table_contexts[CIS Level],"(L2)")'
            next_letter = get_column_letter(column_index_from_string(letter) + 1)
            for cell in range(2,7):
                worksheet[f'{letter}{cell}'].border = BORDER_COMPLETE_BLACK
                worksheet[f'{letter}{cell}'].alignment = ALIGNMENT_HORIZONTAL_CENTER
                worksheet.merge_cells(f'{letter}{cell}:{next_letter}{cell}')
            letter = get_column_letter(column_index_from_string(letter) + 2)
            context_index+=1

    def table_number_workshop(self, worksheet):
        """
        Create a table to count policies by workshop
        """
        worksheet['D8'] = 'Workshops'
        worksheet['D8'].font = FONT_COLOR_BLACK_BOLD
        worksheet['D9'] = 'Remaining policies'
        worksheet['F9'] = '=150-SUM(F10:F17)'
        worksheet['D18'] = 'Total*'
        worksheet['D19'] = 'Total (unique)'
        worksheet['F18'] = '=SUM(F10:F17)'
        worksheet['F19'] = '=COUNTIF(table_contexts[Workshops],"<>_")'
        for cell in range(10, 18):
            worksheet['D' + str(cell)] = 'Workshop-' + str(cell-9)
            worksheet['F' + str(cell)] =\
                '=COUNTIF(table_contexts[Workshops],CONCATENATE("*",D' + str(cell) + ',"*"))'
        for cell in range(9, 18):
            worksheet['D' + str(cell)].border = BORDER_COMPLETE_BLACK
            worksheet['D' + str(cell)].alignment = ALIGNMENT_HORIZONTAL_CENTER
            worksheet['F' + str(cell)].border = BORDER_COMPLETE_BLACK
            worksheet.merge_cells('D' + str(cell) + ':E' + str(cell))
        for cell in range(18, 20):
            worksheet['D' + str(cell)].font = FONT_COLOR_BLACK_BOLD
            worksheet['F' + str(cell)].font = FONT_COLOR_BLACK_BOLD
        worksheet['D21'] =\
            '*Include duplicate of policies if they have been adressed during multiple workshops.'
        worksheet['D21'].alignment = Alignment(vertical='top', wrapText=True)
        worksheet.merge_cells('D21:F24')

    def pie_chart_policies_by_workshop_on_main_dashboard(self, worksheet):
        """
        Create a pie chart to know the pourcentage of policies by workshop
        """
        pie = PieChart()
        labels = Reference(worksheet, min_col=4, max_col=4, min_row=10, max_row=17)
        data = Reference(worksheet, min_col=6, max_col=6, min_row=10, max_row=17)
        pie.add_data(data, titles_from_data=False)
        pie.set_categories(labels)
        pie.title = "Policies by workshop"
        pie.style = 2
        _from = AnchorMarker(
            col = 10,
            row = 2
        )
        to_end = AnchorMarker(
            col = 15,
            row = 19
        )
        pie.anchor = TwoCellAnchor(editAs='twoCell', _from=_from, to=to_end)
        worksheet.add_chart(pie)

    def chart_context(self, worksheet, context, context_index,
                  start_row, cell_nb_tot_policies, cell_nb_tot_l1):
        """
        Create all charts and tables for each context
        """
        worksheet[f'B{start_row}'] = f'CONTEXT {context_index} - {context["Name"]} - Global Data'
        worksheet[f'B{start_row}'].font = FONT_COLOR_BLACK_BOLD
        worksheet[f'B{start_row}'].border = BORDER_COMPLETE_BLACK
        worksheet[f'B{start_row}'].alignment = ALIGNMENT_CENTER_CENTER
        worksheet.merge_cells(f'B{start_row}:N{start_row+2}')
        worksheet[f'F{start_row+5}'] = 'OS'
        worksheet[f'F{start_row+5}'].font = FONT_COLOR_WHITE_BOLD
        worksheet[f'F{start_row+5}'].alignment = ALIGNMENT_HORIZONTAL_CENTER
        worksheet[f'F{start_row+5}'].fill = GREEN_FILL
        worksheet.merge_cells(f'F{start_row+5}:J{start_row+5}')
        worksheet[f'F{start_row+6}'] = 'OS Version'
        worksheet[f'F{start_row+7}'] = 'OS Subversion'
        worksheet[f'F{start_row+7}'].alignment = ALIGNMENT_WRAP_TEXT
        worksheet[f'G{start_row+6}'] = f"='{context_index}-{context['Name']}-log-reported'!$A$10"
        worksheet[f'G{start_row+7}'] = f"='{context_index}-{context['Name']}-log-reported'!$A$11"
        for row in range(start_row+6, start_row+8):
            worksheet[f'F{row}'].font = Font(color='FFFFFF')
            worksheet[f'F{row}'].fill = PatternFill(fill_type='solid', fgColor='808080')
            worksheet[f'F{row}'].border = BORDER_COMPLETE_WHITE
            worksheet[f'G{row}'].alignment = ALIGNMENT_CENTER_CENTER
            worksheet.merge_cells(f'G{row}:J{row}')
        worksheet[f'F{start_row+8}'] = 'Network'
        worksheet[f'F{start_row+8}'].font = FONT_COLOR_WHITE_BOLD
        worksheet[f'F{start_row+8}'].alignment = ALIGNMENT_HORIZONTAL_CENTER
        worksheet[f'F{start_row+8}'].fill = GREEN_FILL
        worksheet.merge_cells(f'F{start_row+8}:J{start_row+8}')
        worksheet[f'F{start_row+9}'] = 'Hostname'
        worksheet[f'F{start_row+10}'] = 'Domain'
        worksheet[f'F{start_row+11}'] = 'Role'
        worksheet[f'G{start_row+9}'] = f"='{context_index}-{context['Name']}-log-reported'!$A$3"
        worksheet[f'G{start_row+10}'] = f"='{context_index}-{context['Name']}-log-reported'!$A$4"
        worksheet[f'G{start_row+11}'] = f"='{context_index}-{context['Name']}-log-reported'!$A$5"
        for row in range(start_row+9, start_row+12):
            worksheet[f'F{row}'].font = Font(color='FFFFFF')
            worksheet[f'F{row}'].fill = PatternFill(fill_type='solid', fgColor='808080')
            worksheet[f'G{row}'].alignment = ALIGNMENT_HORIZONTAL_CENTER
            worksheet[f'F{row}'].border = BORDER_COMPLETE_WHITE
            worksheet.merge_cells(f'G{row}:J{row}')
        start_row_sub_category = start_row + 14
        self.configuration_context(worksheet, start_row_sub_category, context_index, cell_nb_tot_policies)
        self.status_details_context(worksheet, start_row_sub_category, context_index)
        self.configuration_l1_context(worksheet, start_row_sub_category, context_index, cell_nb_tot_l1)

    def configuration_context(self, worksheet, start_row, num_context, cell_nb_tot_policies):
        """
        Create Configuration chart for each context
        """
        worksheet[f'C{start_row}'] = 'Configuration'
        worksheet[f'C{start_row}'].font = FONT_COLOR_WHITE_BOLD
        worksheet[f'C{start_row}'].fill = GREEN_FILL
        worksheet[f'C{start_row}'].alignment = ALIGNMENT_CENTER_CENTER
        worksheet.merge_cells(f'C{start_row}' + ':M' + str(start_row+1))
        worksheet[f'D{start_row+3}'] = 'Before Workshops'
        worksheet[f'J{start_row+3}'] = 'After Workshops'
        worksheet[f'D{start_row+3}'].alignment = ALIGNMENT_HORIZONTAL_CENTER
        worksheet[f'J{start_row+3}'].alignment = ALIGNMENT_HORIZONTAL_CENTER
        worksheet.merge_cells(f'D{start_row+3}:F{start_row+3}')
        worksheet.merge_cells(f'J{start_row+3}:L{start_row+3}')
        worksheet[f'D{start_row+4}'] = 'Passed'
        worksheet[f'J{start_row+4}'] = 'Passed'
        worksheet[f'D{start_row+5}'] = 'High'
        worksheet[f'J{start_row+5}'] = 'High'
        worksheet[f'D{start_row+6}'] = 'Medium'
        worksheet[f'J{start_row+6}'] = 'Medium'
        worksheet[f'D{start_row+7}'] = 'Low'
        worksheet[f'J{start_row+7}'] = 'Low'
        worksheet[f'D{start_row+8}'] = 'Not Applicable'
        worksheet[f'J{start_row+8}'] = 'Not Applicable'
        worksheet[f'D{start_row+9}'] = 'Default Values'
        worksheet[f'J{start_row+9}'] = 'To check'
        worksheet[f'D{start_row+4}'].fill = PatternFill(fill_type='solid', fgColor='00B050')
        worksheet[f'J{start_row+4}'].fill = PatternFill(fill_type='solid', fgColor='00B050')
        worksheet[f'D{start_row+5}'].fill = PatternFill(fill_type='solid', fgColor='C00000')
        worksheet[f'J{start_row+5}'].fill = PatternFill(fill_type='solid', fgColor='C00000')
        worksheet[f'D{start_row+6}'].fill = PatternFill(fill_type='solid', fgColor='FFDA31')
        worksheet[f'J{start_row+6}'].fill = PatternFill(fill_type='solid', fgColor='FFDA31')
        worksheet[f'D{start_row+7}'].fill = PatternFill(fill_type='solid', fgColor='00A2FF')
        worksheet[f'J{start_row+7}'].fill = PatternFill(fill_type='solid', fgColor='00A2FF')
        worksheet[f'D{start_row+8}'].fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
        worksheet[f'J{start_row+8}'].fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
        worksheet[f'D{start_row+8}'].alignment = ALIGNMENT_WRAP_TEXT
        worksheet[f'J{start_row+8}'].alignment = ALIGNMENT_WRAP_TEXT
        worksheet[f'D{start_row+9}'].fill = PatternFill(fill_type='solid', fgColor='808080')
        worksheet[f'J{start_row+9}'].fill = PatternFill(fill_type='solid', fgColor='FFA395')
        worksheet[f'D{start_row+9}'].alignment = ALIGNMENT_WRAP_TEXT
        worksheet[f'E{start_row+9}'] =\
            f'=COUNTIFS(table_contexts[Context{num_context} - Result],"-NODATA-",table_contexts[Context{num_context}],TRUE)'
        worksheet[f'K{start_row+9}'] =\
           f'=COUNTIFS(table_contexts[Context{num_context} - Fixed Value],"to check",table_contexts[Context{num_context}],TRUE)'
        for cell in range(start_row+4, start_row+6):
            worksheet[f'D{cell}'].font = FONT_COLOR_WHITE
            worksheet[f'J{cell}'].font = FONT_COLOR_WHITE
        for cell in range(start_row+4, start_row+9):
            worksheet[f'E{cell}'] =\
                f'=COUNTIFS(table_contexts[Context{num_context} - Computed Severity],D{cell},table_contexts[Context{num_context}],TRUE)'
            worksheet[f'K{cell}'] =\
                f'=COUNTIFS(table_contexts[Context{num_context} - isRecValue],J{cell},table_contexts[Context{num_context}],TRUE)'
        for cell in range(start_row+4, start_row+10):
            worksheet[f'D{cell}'].border = BORDER_COMPLETE_WHITE
            worksheet[f'J{cell}'].border = BORDER_COMPLETE_WHITE
            worksheet[f'E{cell}'].font = FONT_COLOR_BLACK_BOLD
            worksheet[f'K{cell}'].font = FONT_COLOR_BLACK_BOLD
            worksheet[f'F{cell}'] = f'=E{cell}/{cell_nb_tot_policies}'
            worksheet[f'L{cell}'] = f'=K{cell}/{cell_nb_tot_policies}'
            worksheet[f'F{cell}'].number_format = '0%'
            worksheet[f'L{cell}'].number_format = '0%'
        start_row_chart = start_row + 10
        self.pie_chart_sub_category_configuration(worksheet, start_row_chart, 2)

    def pie_chart_sub_category_configuration(self, worksheet, start_row_chart, col_start_char):
        """
        Create 2 pie chart to compare before and after workshops result
        """
        pie_before = PieChart()
        labels = Reference(worksheet, min_col=col_start_char+2, max_col=col_start_char+2,
                        min_row=start_row_chart-6, max_row=start_row_chart-2)
        data = Reference(worksheet, min_col=col_start_char+3, max_col=col_start_char+3,
                        min_row=start_row_chart-6, max_row=start_row_chart-2)
        pie_before.add_data(data, titles_from_data=False)
        pie_before.set_categories(labels)
        pie_before.style = 2
        serie = pie_before.series[0]
        for part, colors in enumerate(['00B050', 'C00000', 'FFDA31', '00A2FF', 'D9D9D9']):
            data_part = DataPoint(idx=part)
            data_part.graphicalProperties.solidFill = colors
            serie.dPt.append(data_part)
        _from = AnchorMarker(
            col = col_start_char,
            row = start_row_chart
        )
        to_end = AnchorMarker(
            col = col_start_char+5,
            row = start_row_chart+12
        )
        pie_before.anchor = TwoCellAnchor(editAs='twoCell', _from=_from, to=to_end)
        worksheet.add_chart(pie_before)
        pie_after = PieChart()
        labels = Reference(worksheet, min_col=col_start_char+8, max_col=col_start_char+8,
                        min_row=start_row_chart-6, max_row=start_row_chart-2)
        data = Reference(worksheet, min_col=col_start_char+9, max_col=col_start_char+9,
                        min_row=start_row_chart-6, max_row=start_row_chart-2)
        pie_after.add_data(data, titles_from_data=False)
        pie_after.set_categories(labels)
        pie_after.style = 2
        serie = pie_after.series[0]
        for part, colors in enumerate(['00B050', 'C00000', 'FFDA31', '00A2FF', 'D9D9D9']):
            data_part = DataPoint(idx=part)
            data_part.graphicalProperties.solidFill = colors
            serie.dPt.append(data_part)
        _from = AnchorMarker(
            col = col_start_char+6,
            row = start_row_chart
        )
        to_end = AnchorMarker(
            col = col_start_char+11,
            row = start_row_chart+12
        )
        pie_after.anchor = TwoCellAnchor(editAs='twoCell', _from=_from, to=to_end)
        worksheet.add_chart(pie_after)

    def status_details_context(self, worksheet, start_row, num_context):
        """
        Create Status - Details chart for each context
        """
        worksheet['O' + str(start_row)] = 'Status - Details'
        worksheet['O' + str(start_row)].font = FONT_COLOR_WHITE_BOLD
        worksheet['O' + str(start_row)].fill = GREEN_FILL
        worksheet['O' + str(start_row)].alignment = ALIGNMENT_CENTER_CENTER
        worksheet.merge_cells('O' + str(start_row) + ':AD' + str(start_row+1))
        number_category = pd.read_excel(self.path, sheet_name='All-Policies')['Category'].nunique() + 3
        worksheet['O' + str(start_row+4)] = ArrayFormula(ref='O' + str(start_row+4) + ':O' +
                                                        str(start_row+number_category),
                                                        text='=_xlfn.UNIQUE(table_contexts[Category])')
        worksheet['P' + str(start_row+3)] = 'High'
        worksheet['Q' + str(start_row+3)] = 'Medium'
        worksheet['R' + str(start_row+3)] = 'Low'
        worksheet['S' + str(start_row+3)] = 'Passed'
        worksheet['T' + str(start_row+3)] = 'Total'
        for column in range(16, 21):
            letter = get_column_letter(column)
            worksheet[letter + str(start_row+3)].font = FONT_COLOR_BLACK_BOLD
            worksheet[letter + str(start_row+3)].border = BORDER_COMPLETE_BLACK
        for cell in range(start_row+3, start_row+number_category+1):
            worksheet['O' + str(cell)].border = BORDER_COMPLETE_BLACK
            worksheet['O' + str(cell)].font = FONT_COLOR_BLACK_BOLD
        for cell in range(start_row+4, start_row+number_category+1):
            worksheet['P' + str(cell)] = f'=COUNTIFS(table_contexts[Category],$O{cell},table_contexts[Context{num_context} - isRecValue],$P${start_row+3})/$T{cell}'
            worksheet['T' + str(cell)] = '=COUNTIFS(table_contexts[Category],$O' + str(cell) + ')'
            for column in range(17, 20):
                letter = get_column_letter(column)
                worksheet[letter + str(cell)] =\
                    f'=COUNTIFS(table_contexts[Category],$O{cell},table_contexts[Context{num_context} - isRecValue],{letter}${start_row+3})/$T{cell}'
            for column in range(16, 20):
                letter = get_column_letter(column)
                worksheet[letter + str(cell)].number_format = '0%'
            for column in range(16, 21):
                letter = get_column_letter(column)
                worksheet[letter + str(cell)].border = BORDER_COMPLETE_BLACK
        worksheet.conditional_formatting.add('S' + str(start_row+4) + ':S' +
                                            str(start_row+number_category),COLORSCALE)
        start_row_chart = start_row + 3
        self.bar_chart_status_details(worksheet, start_row_chart, number_category)

    def bar_chart_status_details(self, worksheet, start_row_chart, number_category):
        """
        Create a percent stacked chart of the repartition status by categories
        """
        bar_chart = BarChart()
        bar_chart.type = 'bar'
        bar_chart.style = 2
        bar_chart.title = 'Repartition status by categories'
        labels = Reference(worksheet, min_col=15, max_col=15,
                        min_row=start_row_chart+1, max_row=start_row_chart+number_category-3)
        data = Reference(worksheet, min_col=16, max_col=19,
                        min_row=start_row_chart, max_row=start_row_chart+number_category-3)
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(labels)
        bar_chart.grouping = "percentStacked"
        bar_chart.overlap = 100
        colors = ['C00000', 'FFDA31', '00A2FF', '00B050']
        for color, serie in enumerate(bar_chart.series):
            serie.graphicalProperties.solidFill = ColorChoice(srgbClr=colors[color])
        _from = AnchorMarker(
            col = 21,
            row = start_row_chart-1
        )
        to_end = AnchorMarker(
            col = 30,
            row = start_row_chart+number_category-3
        )
        bar_chart.anchor = TwoCellAnchor(editAs='twoCell', _from=_from, to=to_end)
        worksheet.add_chart(bar_chart)

    def configuration_l1_context(self, worksheet, start_row, num_context, cell_nb_tot_l1):
        """
        Create Configuration chart for each context
        """
        worksheet['AF' + str(start_row)] = 'Configuration (L1)'
        worksheet['AF' + str(start_row)].font = FONT_COLOR_WHITE_BOLD
        worksheet['AF' + str(start_row)].fill = GREEN_FILL
        worksheet['AF' + str(start_row)].alignment = ALIGNMENT_CENTER_CENTER
        worksheet.merge_cells('AF' + str(start_row) + ':AV' + str(start_row+1))
        worksheet['AG' + str(start_row+3)] = 'Before Workshops'
        worksheet['AM' + str(start_row+3)] = 'After Workshops'
        worksheet['AS' + str(start_row+3)] = 'To check'
        worksheet['AG' + str(start_row+3)].alignment = ALIGNMENT_HORIZONTAL_CENTER
        worksheet['AM' + str(start_row+3)].alignment = ALIGNMENT_HORIZONTAL_CENTER
        worksheet['AS' + str(start_row+3)].alignment = ALIGNMENT_HORIZONTAL_CENTER
        worksheet['AS' + str(start_row+3)].fill = PatternFill(fill_type='solid', fgColor='FFA395')
        worksheet.merge_cells('AG' + str(start_row+3) + ':AI' + str(start_row+3))
        worksheet.merge_cells('AM' + str(start_row+3) + ':AO' + str(start_row+3))
        worksheet.merge_cells('AS' + str(start_row+3) + ':AT' + str(start_row+3))
        worksheet['AG' + str(start_row+4)] = 'Passed'
        worksheet['AM' + str(start_row+4)] = 'Passed'
        worksheet['AS' + str(start_row+4)] = 'Passed'
        worksheet['AG' + str(start_row+5)] = 'High'
        worksheet['AM' + str(start_row+5)] = 'High'
        worksheet['AS' + str(start_row+5)] = 'High'
        worksheet['AG' + str(start_row+6)] = 'Medium'
        worksheet['AM' + str(start_row+6)] = 'Medium'
        worksheet['AS' + str(start_row+6)] = 'Medium'
        worksheet['AG' + str(start_row+7)] = 'Low'
        worksheet['AM' + str(start_row+7)] = 'Low'
        worksheet['AS' + str(start_row+7)] = 'Low'
        worksheet['AG' + str(start_row+8)] = 'Not Applicable'
        worksheet['AM' + str(start_row+8)] = 'Not Applicable'
        worksheet['AS' + str(start_row+8)] = 'Total'
        worksheet['AG' + str(start_row+4)].fill = PatternFill(fill_type='solid', fgColor='00B050')
        worksheet['AM' + str(start_row+4)].fill = PatternFill(fill_type='solid', fgColor='00B050')
        worksheet['AS' + str(start_row+4)].fill = PatternFill(fill_type='solid', fgColor='00B050')
        worksheet['AG' + str(start_row+5)].fill = PatternFill(fill_type='solid', fgColor='C00000')
        worksheet['AM' + str(start_row+5)].fill = PatternFill(fill_type='solid', fgColor='C00000')
        worksheet['AS' + str(start_row+5)].fill = PatternFill(fill_type='solid', fgColor='C00000')
        worksheet['AG' + str(start_row+6)].fill = PatternFill(fill_type='solid', fgColor='FFDA31')
        worksheet['AM' + str(start_row+6)].fill = PatternFill(fill_type='solid', fgColor='FFDA31')
        worksheet['AS' + str(start_row+6)].fill = PatternFill(fill_type='solid', fgColor='FFDA31')
        worksheet['AG' + str(start_row+7)].fill = PatternFill(fill_type='solid', fgColor='00A2FF')
        worksheet['AM' + str(start_row+7)].fill = PatternFill(fill_type='solid', fgColor='00A2FF')
        worksheet['AS' + str(start_row+7)].fill = PatternFill(fill_type='solid', fgColor='00A2FF')
        worksheet['AG' + str(start_row+8)].fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
        worksheet['AM' + str(start_row+8)].fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
        worksheet['AS' + str(start_row+8)].fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
        worksheet['AG' + str(start_row+8)].alignment = ALIGNMENT_WRAP_TEXT
        worksheet['AM' + str(start_row+8)].alignment = ALIGNMENT_WRAP_TEXT
        worksheet['AT' + str(start_row+8)] =\
            '=SUM(AT' + str(start_row+4) + ':AT' + str(start_row+7) + ')'
        for cell in range(start_row+4, start_row+6):
            worksheet['AG' + str(cell)].font = FONT_COLOR_WHITE
            worksheet['AM' + str(cell)].font = FONT_COLOR_WHITE
        for cell in range(start_row+4, start_row+8):
            worksheet['AT' + str(cell)] =\
                '=COUNTIFS(table_contexts[Context' + str(num_context) + ' - isRecValue],AS' + str(cell) + ',table_contexts[Context' + str(num_context) + '],TRUE,table_contexts[CIS Level],"(L1)",table_contexts[Context' + str(num_context) + ' - Fixed Value],"to check")'
        for cell in range(start_row+4, start_row+9):
            worksheet['AH' + str(cell)] =\
                '=COUNTIFS(table_contexts[Context' + str(num_context) + ' - Computed Severity],AG' + str(cell) + ',table_contexts[Context' + str(num_context) + '],TRUE,table_contexts[CIS Level],"(L1)")'
            worksheet['AN' + str(cell)] =\
                '=COUNTIFS(table_contexts[Context' + str(num_context) + ' - isRecValue],AM' + str(cell) + ',table_contexts[Context' + str(num_context) + '],TRUE,table_contexts[CIS Level],"(L1)")'
            worksheet['AG' + str(cell)].border = BORDER_COMPLETE_WHITE
            worksheet['AM' + str(cell)].border = BORDER_COMPLETE_WHITE
            worksheet['AH' + str(cell)].font = FONT_COLOR_BLACK_BOLD
            worksheet['AN' + str(cell)].font = FONT_COLOR_BLACK_BOLD
            worksheet['AI' + str(cell)] = '=AH' + str(cell) + '/' + str(cell_nb_tot_l1)
            worksheet['AO' + str(cell)] = '=AN' + str(cell) + '/' + str(cell_nb_tot_l1)
            worksheet['AI' + str(cell)].number_format = '0%'
            worksheet['AO' + str(cell)].number_format = '0%'
        start_row_chart = start_row + 9
        self.pie_chart_sub_category_l1(worksheet, start_row_chart, 31)

    def pie_chart_sub_category_l1(self, worksheet, start_row_chart, col_start_char):
        """
        Create tge last chart to know the repartition of to check by status
        """
        pie_before = PieChart()
        labels = Reference(worksheet, min_col=col_start_char+2, max_col=col_start_char+2,
                        min_row=start_row_chart-5, max_row=start_row_chart-1)
        data = Reference(worksheet, min_col=col_start_char+3, max_col=col_start_char+3,
                        min_row=start_row_chart-5, max_row=start_row_chart-1)
        pie_before.add_data(data, titles_from_data=False)
        pie_before.set_categories(labels)
        serie = pie_before.series[0]
        for part, colors in enumerate(['00B050', 'C00000', 'FFDA31', '00A2FF', 'D9D9D9']):
            data_part = DataPoint(idx=part)
            data_part.graphicalProperties.solidFill = colors
            serie.dPt.append(data_part)
        pie_before.style = 2
        _from = AnchorMarker(
            col = col_start_char,
            row = start_row_chart
        )
        to_end = AnchorMarker(
            col = col_start_char+5,
            row = start_row_chart+13
        )
        pie_before.anchor = TwoCellAnchor(editAs='twoCell', _from=_from, to=to_end)
        worksheet.add_chart(pie_before)
        pie_after = PieChart()
        labels = Reference(worksheet, min_col=col_start_char+8, max_col=col_start_char+8,
                        min_row=start_row_chart-5, max_row=start_row_chart-1)
        data = Reference(worksheet, min_col=col_start_char+9, max_col=col_start_char+9,
                        min_row=start_row_chart-5, max_row=start_row_chart-1)
        pie_after.add_data(data, titles_from_data=False)
        pie_after.set_categories(labels)
        serie = pie_after.series[0]
        for part, colors in enumerate(['00B050', 'C00000', 'FFDA31', '00A2FF', 'D9D9D9']):
            data_part = DataPoint(idx=part)
            data_part.graphicalProperties.solidFill = colors
            serie.dPt.append(data_part)
        pie_after.style = 2
        _from = AnchorMarker(
            col = col_start_char+6,
            row = start_row_chart
        )
        to_end = AnchorMarker(
            col = col_start_char+11,
            row = start_row_chart+13
        )
        pie_after.anchor = TwoCellAnchor(editAs='twoCell', _from=_from, to=to_end)
        worksheet.add_chart(pie_after)
        pie_to_check = PieChart()
        labels = Reference(worksheet, min_col=col_start_char+14, max_col=col_start_char+14,
                        min_row=start_row_chart-5, max_row=start_row_chart-2)
        data = Reference(worksheet, min_col=col_start_char+15, max_col=col_start_char+15,
                        min_row=start_row_chart-5, max_row=start_row_chart-2)
        pie_to_check.add_data(data, titles_from_data=False)
        pie_to_check.set_categories(labels)
        serie = pie_to_check.series[0]
        for part, colors in enumerate(['00B050', 'C00000', 'FFDA31', '00A2FF']):
            data_part = DataPoint(idx=part)
            data_part.graphicalProperties.solidFill = colors
            serie.dPt.append(data_part)
        pie_to_check.style = 2
        _from = AnchorMarker(
            col = col_start_char+12,
            row = start_row_chart
        )
        to_end = AnchorMarker(
            col = col_start_char+17,
            row = start_row_chart+13
        )
        pie_to_check.anchor = TwoCellAnchor(editAs='twoCell', _from=_from, to=to_end)
        worksheet.add_chart(pie_to_check)

    def construct_workshop_dashboard_sheet(self):
            """
            Launch all functions to create the entire sheet
            """
            workbook = load_workbook(self.path, data_only=False)
            worksheet = workbook['Dashboards - Workshops']
            for columns in range(1, 71):
                worksheet.column_dimensions[get_column_letter(columns)].width = 11
            worksheet.column_dimensions['B'].width = 19.5
            self.policies_by_workshop(worksheet)
            self.summary_before_after_workshops(worksheet)
            self.construct_each_workshop_chart(worksheet)
            workbook.save(self.path)

    def policies_by_workshop(self, worksheet):
        """
        Create table + chart about policies by workshop
        """
        worksheet['B2'] = 'Number of Policies'
        worksheet['B3'] = 'Number of Categories'
        worksheet['C2'] = '=COUNTIF(All_policies[ID],"<>")'
        worksheet['C3'] = '=COUNTA(_xlfn.UNIQUE(All_policies[Category]))'
        worksheet['B5'] = 'Workshops'
        worksheet['B5'].font = FONT_COLOR_BLACK_BOLD
        worksheet['B6'] = 'Remaining policies'
        worksheet['C6'] = '=150-C15'
        worksheet['B15'] = 'Total'
        worksheet['C15'] = '=SUM(C7:C14)'
        for cell in range(7, 15):
            worksheet['B' + str(cell)] = 'Workshop-' + str(cell-6)
            worksheet['C' + str(cell)] =\
                '=COUNTIF(table_contexts[Workshops],CONCATENATE("*",B' + str(cell) + ',"*"))'
        self.pie_chart_policies_by_workshop_on_workshop_dashboard(worksheet)

    def pie_chart_policies_by_workshop_on_workshop_dashboard(self, worksheet):
        """
        Create a pie chart to know the pourcentage of policies by workshop
        """
        pie = PieChart()
        labels = Reference(worksheet, min_col=2, max_col=2, min_row=7, max_row=14)
        data = Reference(worksheet, min_col=3, max_col=3, min_row=7, max_row=14)
        pie.add_data(data, titles_from_data=False)
        pie.set_categories(labels)
        pie.title = "Policies by workshop"
        pie.style = 2
        _from = AnchorMarker(
            col = 4,
            row = 2
        )
        to_end = AnchorMarker(
            col = 13,
            row = 20
        )
        pie.anchor = TwoCellAnchor(editAs='twoCell', _from=_from, to=to_end)
        worksheet.add_chart(pie)

    def summary_before_after_workshops(self, worksheet):
        """
        Create table and charts about the summary of all workshops
        """
        worksheet['F24'] = 'Before workshops'
        worksheet['J24'] = 'After workshops'
        worksheet['F24'].alignment = ALIGNMENT_HORIZONTAL_CENTER
        worksheet['J24'].alignment = ALIGNMENT_HORIZONTAL_CENTER
        worksheet['F24'].border = Border(top=SIDE_BLACK_MEDIUM,
                                        left=SIDE_BLACK_MEDIUM,
                                        bottom=SIDE_BLACK_THIN)
        worksheet['J24'].border = Border(top=SIDE_BLACK_MEDIUM,
                                        right=SIDE_BLACK_MEDIUM,
                                        bottom=SIDE_BLACK_THIN)
        worksheet['I24'].border = Border(top=SIDE_BLACK_MEDIUM,
                                        bottom=SIDE_BLACK_THIN)
        worksheet.merge_cells('F24:H24')
        worksheet.merge_cells('J24:L24')

        context_index = 1
        starting_row = 25
        for context in self.contexts:
            self.create_context_summary(worksheet, context, 1, 25)
            context_index+=1
            starting_row+=6

    def create_context_summary(self, worksheet, context, num_context, row_start):
        """
        Create the summary of workshops by context
        """
        worksheet['F' + str(row_start)] = context['Name']
        worksheet['F' + str(row_start)].font = FONT_COLOR_BLACK_BOLD
        worksheet['F' + str(row_start)].alignment = ALIGNMENT_HORIZONTAL_CENTER
        worksheet['F' + str(row_start)].border = Border(left=SIDE_BLACK_MEDIUM,
                                                        right=SIDE_BLACK_MEDIUM)
        worksheet.merge_cells('F' + str(row_start) + ':L' + str(row_start))
        worksheet['F' + str(row_start+1)] = 'Passed'
        worksheet['J' + str(row_start+1)] = 'Passed'
        worksheet['F' + str(row_start+2)] = 'High'
        worksheet['J' + str(row_start+2)] = 'High'
        worksheet['F' + str(row_start+3)] = 'Medium'
        worksheet['J' + str(row_start+3)] = 'Medium'
        worksheet['F' + str(row_start+4)] = 'Low'
        worksheet['J' + str(row_start+4)] = 'Low'
        worksheet['F' + str(row_start+1)].fill = PatternFill(fill_type='solid', fgColor='00B050')
        worksheet['F' + str(row_start+1)].border = Border(left=SIDE_BLACK_MEDIUM)
        worksheet['J' + str(row_start+1)].fill = PatternFill(fill_type='solid', fgColor='00B050')
        worksheet['F' + str(row_start+2)].fill = PatternFill(fill_type='solid', fgColor='C00000')
        worksheet['F' + str(row_start+2)].border = Border(left=SIDE_BLACK_MEDIUM)
        worksheet['J' + str(row_start+2)].fill = PatternFill(fill_type='solid', fgColor='C00000')
        worksheet['F' + str(row_start+3)].fill = PatternFill(fill_type='solid', fgColor='FFDA31')
        worksheet['F' + str(row_start+3)].border = Border(left=SIDE_BLACK_MEDIUM)
        worksheet['J' + str(row_start+3)].fill = PatternFill(fill_type='solid', fgColor='FFDA31')
        worksheet['F' + str(row_start+4)].fill = PatternFill(fill_type='solid', fgColor='00A2FF')
        worksheet['F' + str(row_start+4)].border = Border(left=SIDE_BLACK_MEDIUM)
        worksheet['J' + str(row_start+4)].fill = PatternFill(fill_type='solid', fgColor='00A2FF')
        for cell in range(row_start+1, row_start+3):
            worksheet['F' + str(cell)].font = FONT_COLOR_WHITE
            worksheet['J' + str(cell)].font = FONT_COLOR_WHITE
        for cell in range(row_start+1, row_start+5):
            if num_context == 1:
                worksheet['G' + str(cell)] = '=Dashboards!E' + str(cell+19)
                worksheet['H' + str(cell)] = '=Dashboards!F' + str(cell+19)
                worksheet['K' + str(cell)] = '=Dashboards!K' + str(cell+19)
                worksheet['L' + str(cell)] = '=Dashboards!L' + str(cell+19)
            elif num_context == 2:
                worksheet['G' + str(cell)] = '=Dashboards!E' + str(cell+55)
                worksheet['H' + str(cell)] = '=Dashboards!F' + str(cell+55)
                worksheet['K' + str(cell)] = '=Dashboards!K' + str(cell+55)
                worksheet['L' + str(cell)] = '=Dashboards!L' + str(cell+55)
            elif num_context == 3:
                worksheet['G' + str(cell)] = '=Dashboards!E' + str(cell+91)
                worksheet['H' + str(cell)] = '=Dashboards!F' + str(cell+91)
                worksheet['K' + str(cell)] = '=Dashboards!K' + str(cell+91)
                worksheet['L' + str(cell)] = '=Dashboards!L' + str(cell+91)
            worksheet['G' + str(cell)].font = FONT_COLOR_BLACK_BOLD
            worksheet['K' + str(cell)].font = FONT_COLOR_BLACK_BOLD
            worksheet['F' + str(cell)].border = Border(left=SIDE_BLACK_MEDIUM,
                                                    top=SIDE_WHITE_THIN,
                                                    right=SIDE_WHITE_THIN,
                                                    bottom=SIDE_WHITE_THIN)
            worksheet['J' + str(cell)].border = BORDER_COMPLETE_WHITE
            worksheet['L' + str(cell)].border = Border(right=SIDE_BLACK_MEDIUM)
            worksheet['H' + str(cell)].number_format = '0%'
            worksheet['L' + str(cell)].number_format = '0%'
        worksheet['F' + str(row_start+5)].border = Border(left=SIDE_BLACK_MEDIUM,
                                                        bottom=SIDE_BLACK_MEDIUM)
        worksheet['J' + str(row_start+5)] = 'To check'
        worksheet['J' + str(row_start+5)].fill = PatternFill(fill_type='solid', fgColor='FFA395')
        worksheet['J' + str(row_start+5)].border = Border(top=SIDE_WHITE_THIN,
                                                        left=SIDE_WHITE_THIN,
                                                        right=SIDE_WHITE_THIN,
                                                        bottom=SIDE_BLACK_MEDIUM)
        if num_context == 1:
            worksheet['K' + str(row_start+5)] = '=Dashboards!K' + str(row_start+29)
            worksheet['L' + str(row_start+5)] = '=Dashboards!L' + str(row_start+29)
        if num_context == 2:
            worksheet['K' + str(row_start+5)] = '=Dashboards!K' + str(row_start+65)
            worksheet['L' + str(row_start+5)] = '=Dashboards!L' + str(row_start+65)
        if num_context == 3:
            worksheet['K' + str(row_start+5)] = '=Dashboards!K' + str(row_start+101)
            worksheet['L' + str(row_start+5)] = '=Dashboards!L' + str(row_start+101)
        worksheet['K' + str(row_start+5)].font = FONT_COLOR_BLACK_BOLD
        worksheet['K' + str(row_start+5)].border = Border(bottom=SIDE_BLACK_MEDIUM)
        worksheet['L' + str(row_start+5)].number_format = '0%'
        worksheet['L' + str(row_start+5)].border = Border(right=SIDE_BLACK_MEDIUM,
                                                        bottom=SIDE_BLACK_MEDIUM)
        for column in range(7, 10):
            letter = get_column_letter(column)
            worksheet[letter + str(row_start+5)].border = Border(bottom=SIDE_BLACK_MEDIUM)

    def construct_each_workshop_chart(self, worksheet):
        """
        Construct charts for each workshop
        """
        workshops = 0
        for column in range(15, 71, 7):
            workshops += 1
            worksheet[get_column_letter(column) + '2'].border = Border(top=SIDE_BLACK_MEDIUM,
                                                                    left=SIDE_BLACK_MEDIUM)
            worksheet[get_column_letter(column) + '31'].border = Border(bottom=SIDE_BLACK_MEDIUM,
                                                                    left=SIDE_BLACK_MEDIUM)
            worksheet[get_column_letter(column+6) + '2'].border = Border(top=SIDE_BLACK_MEDIUM,
                                                                        right=SIDE_BLACK_MEDIUM)
            worksheet[get_column_letter(column+6) + '31'].border = Border(bottom=SIDE_BLACK_MEDIUM,
                                                                        right=SIDE_BLACK_MEDIUM)
            for cell in range(3, 31):
                worksheet[get_column_letter(column) + str(cell)].border = Border(left=SIDE_BLACK_MEDIUM)
                worksheet[get_column_letter(column+6)
                        + str(cell)].border = Border(right=SIDE_BLACK_MEDIUM)
            for cell in range(column+1, column+6):
                worksheet[get_column_letter(cell) + '2'].border = Border(top=SIDE_BLACK_MEDIUM)
                worksheet[get_column_letter(cell)
                        + '31'].border = Border(bottom=SIDE_BLACK_MEDIUM)
            worksheet[get_column_letter(column+1) + '3'] = 'Workshop' + str(workshops)
            worksheet[get_column_letter(column+1) + '3'].font = FONT_COLOR_BLACK_BOLD
            worksheet[get_column_letter(column+1) + '3'].alignment = ALIGNMENT_HORIZONTAL_CENTER
            worksheet.merge_cells(get_column_letter(column+1) +
                                '3:' + get_column_letter(column+5) + '4')
            worksheet[get_column_letter(column+1) + '6'] = 'Policies'
            worksheet[get_column_letter(column+2) + '6'] =\
                '=COUNTIF(table_contexts[Workshops],CONCATENATE("*","Workshop-' + str(workshops) + '","*"))'

            context_index = 1
            starting_row = 7
            for context in self.contexts:
                self.chart_by_whorkshop_by_context(worksheet, context, 1, column, 7, workshops)
                context_index+=1
                starting_row+=8

    def chart_by_whorkshop_by_context(self, worksheet, context, num_context,
                                    column, start_row, workshops):
        """
        Create table and chart for each context by workshop
        """
        worksheet[get_column_letter(column+1) + str(start_row+1)] = context['Name']
        worksheet[get_column_letter(column+1) + str(start_row+2)] = 'Passed'
        worksheet[get_column_letter(column+1) + str(start_row+3)] = 'High'
        worksheet[get_column_letter(column+1) + str(start_row+4)] = 'Medium'
        worksheet[get_column_letter(column+1) + str(start_row+5)] = 'Low'
        worksheet[get_column_letter(column+1) + str(start_row+2)].fill =\
            PatternFill(fill_type='solid', fgColor='00B050')
        worksheet[get_column_letter(column+1) + str(start_row+3)].fill =\
            PatternFill(fill_type='solid', fgColor='C00000')
        worksheet[get_column_letter(column+1) + str(start_row+4)].fill =\
            PatternFill(fill_type='solid', fgColor='FFDA31')
        worksheet[get_column_letter(column+1) + str(start_row+5)].fill =\
            PatternFill(fill_type='solid', fgColor='00A2FF')
        worksheet[get_column_letter(column+2) + str(start_row+2)] =\
            '=COUNTIFS(table_contexts[Context' + str(num_context) + ' - Computed Severity],' + get_column_letter(column+1) + str(start_row+2) + ',table_contexts[Workshops],"*Workshop-' + str(workshops) + '*",table_contexts[Context' + str(num_context) + '],TRUE)'
        worksheet[get_column_letter(column+2) + str(start_row+3)] =\
            '=COUNTIFS(table_contexts[Context' + str(num_context) + ' - Computed Severity],' + get_column_letter(column+1) + str(start_row+3) + ',table_contexts[Workshops],"*Workshop-' + str(workshops) + '*",table_contexts[Context' + str(num_context) + '],TRUE)'
        worksheet[get_column_letter(column+2) + str(start_row+4)] =\
            '=COUNTIFS(table_contexts[Context' + str(num_context) + ' - Computed Severity],' + get_column_letter(column+1) + str(start_row+4) + ',table_contexts[Workshops],"*Workshop-' + str(workshops) + '*",table_contexts[Context' + str(num_context) + '],TRUE)'
        worksheet[get_column_letter(column+2) + str(start_row+5)] =\
            '=COUNTIFS(table_contexts[Context' + str(num_context) + ' - Computed Severity],' + get_column_letter(column+1) + str(start_row+5) + ',table_contexts[Workshops],"*Workshop-' + str(workshops) + '*",table_contexts[Context' + str(num_context) + '],TRUE)'
        pie = PieChart()
        labels = Reference(worksheet, min_col=column+1,
                        max_col=column+1, min_row=start_row+2, max_row=start_row+5)
        data = Reference(worksheet, min_col=column+2,
                        max_col=column+2, min_row=start_row+2, max_row=start_row+5)
        pie.add_data(data, titles_from_data=False)
        pie.set_categories(labels)
        pie.style = 2
        serie = pie.series[0]
        for part, colors in enumerate(['00B050', 'C00000', 'FFDA31', '00A2FF']):
            data_part = DataPoint(idx=part)
            data_part.graphicalProperties.solidFill = colors
            serie.dPt.append(data_part)
        _from = AnchorMarker(
            col = column+2,
            row = start_row
        )
        to_end = AnchorMarker(
            col = column+5,
            row = start_row+6
        )
        pie.anchor = TwoCellAnchor(editAs='twoCell', _from=_from, to=to_end)
        worksheet.add_chart(pie)
